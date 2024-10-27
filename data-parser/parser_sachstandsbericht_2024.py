from openpyxl import load_workbook
import json
from pathlib import Path

# valid rows start with one of these prefixes
# rows that start with something different will be appended to the previous one
ROW_TITLE_PREFIXES = [
    "Titel",
    "Kurzbeschreibung",
    "Federführung",
    "Sachstand",
    "Plan 2025",
    "Kosten",
    "CO2-Reduktionspo",
    "Indikatoren (Gesamtmaßnahme)",
    "!",  # Wirtschaftsplanung
    "Informationen zu",
    "Haushaltsansätze/ Planung",
    "Teilfinanzplan",
    "Teilergebnisplan",
]


def extract_from_excel_to_json(
    input_path: str = "data/raw/Sachstandbericht_Klima_plus_Haushaltsinformationen/",
    output_path: str = "data/parsed/Sachstandbericht_Klima_plus_Haushaltsinformationen.json",
):
    """
    Extract information from multiple Excel files and convert them into a single JSON file.

    This function iterates through multiple Excel files located in a specified directory,
    extracts relevant data from each of them, and compiles the data into a JSON file.

    The data extraction process involves gathering cell values from each sheet in an Excel file.
    Each row of the sheet is combined to form a title-value pair. The title is formed by combining
    cell values from the first column of a row, and the value is formed by combining cell values from
    the remaining columns. Rows are considered part of the same group if they do not start with
    any of the specified prefixes (ROW_TITLE_PREFIXES) and are appended to the previous row.
    The process continues until all rows in a sheet have been processed.

    The function then proceeds to the next sheet until all sheets in an Excel file have been processed.
    The process repeats for all Excel files in the directory.

    The end result is a JSON file, where each key-value pair corresponds to an Excel filename and its
    extracted data respectively. The extracted data for each file is also a dictionary where each key-value
    pair corresponds to a title and its associated value.

    Note:
    1. The function assumes that the Excel files and resulting JSON file are located in the
    'data/raw/Sachstandbericht_Klima_plus_Haushaltsinformationen/' and
    'data/parsed/Sachstandbericht_Klima_plus_Haushaltsinformationen.json' directories respectively.
    2. The function uses the openpyxl library to interact with Excel files.

    Raises:
        FileNotFoundError: If the specified directory does not contain any Excel files.
        JSONDecodeError: If there is an error in encoding the final results into a JSON file.
    """
    results = {}

    # iterate through the excel files
    for excel_path in Path(input_path).glob("*.xlsx"):
        if excel_path.name.startswith("~$"):
            continue  # temporary file

        # open the file
        wb = load_workbook(excel_path)

        # collect results for this file here
        results_file = {}

        # go through the sheets (one measure per sheet)
        for sheetname in wb.sheetnames:
            result_sheet = {}
            ws = wb[sheetname]

            current_row_start = 1
            current_row_end = 1

            # first we check how many excel rows the current table row has by going down until we hit one of the valid prefixes
            while current_row_start < ws.max_row:
                while current_row_end < ws.max_row:
                    next_row = current_row_end + 1
                    if not any(
                        [
                            ws.cell(next_row, 1).value is not None
                            and ws.cell(next_row, 1).value.startswith(prefix)
                            for prefix in ROW_TITLE_PREFIXES
                        ]
                    ):
                        # belongs to former row
                        current_row_end = next_row
                    else:
                        break

                # join together the title
                title = "".join(
                    [
                        ws.cell(row, 1).value
                        for row in range(current_row_start, current_row_end + 1)
                        if ws.cell(row, 1).value is not None
                    ]
                ).strip()

                # join together the value
                value = ""
                for row in range(current_row_start, current_row_end + 1):
                    for col in range(2, ws.max_column + 1):
                        if ws.cell(row, col).value is not None:
                            value += str(ws.cell(row, col).value)

                    value += "\n"

                # remove unneeded whitespaces and newlines
                value = value.strip()

                result_sheet[title] = value

                # go to the next row
                current_row_start = current_row_end + 1
                current_row_end = current_row_start

            results_file[result_sheet["Titel"]] = result_sheet

        results[excel_path.stem] = results_file

    with open(
        output_path,
        "w",
        encoding="utf-8",
    ) as fp:
        json.dump(results, fp, ensure_ascii=False, indent=2)


def fill_implementations_from_additional_data(
    additional_data_path: str = "data/additional_data.json",
    implementations_path: str = "data/implementations.json",
):
    with open(implementations_path, "r", encoding="utf-8") as fp:
        implementations = json.load(fp=fp)

    implementations_map = {impl["id"]: impl for impl in implementations}

    with open(additional_data_path, "r", encoding="utf-8") as fp:
        additional_data = json.load(fp=fp)

    for entry in additional_data:
        if "state" not in entry:
            continue

        state = entry["state"]

        if "date" not in state:
            continue

        sachstand_date = state["date"]
        measure_id = entry["id"]

        implementation = implementations_map.get(measure_id, None)

        if implementation is None:
            new_implementation = {
                "id": measure_id,
                "type": "binary",
                "values": [
                    {"value": "in_progress", "date": sachstand_date},
                ],
            }

            implementations_map[measure_id] = new_implementation
            implementations.append(new_implementation)

        else:
            if implementation["type"] != "binary":
                print(f"Skipping {measure_id}: type != binary")
                continue

            implementation_values = implementation["values"]
            if (
                implementation_values[-1]["value"] == "completed"
                or implementation_values[-1]["date"] == sachstand_date
            ):
                print(f"Skipping {measure_id}: value != unknown")
                continue

            implementation_values.append(
                {"value": "in_progress", "date": sachstand_date}
            )

    with open(implementations_path, "w", encoding="utf-8") as fp:
        json.dump(implementations, fp=fp, indent=2, ensure_ascii=False)


def find_unmatched_entries(
    additional_data_path: str = "data/additional_data.json",
    sachstand_path: str = "data/parsed/Sachstandbericht_Klima_plus_Haushaltsinformationen.json",
):
    with open(sachstand_path, "r", encoding="utf-8") as fp:
        sachstand = json.load(fp=fp)

    sachstand_keys = set()

    for area, area_dict in sachstand.items():
        for key, _ in area_dict.items():
            sachstand_keys.add(key)

    with open(additional_data_path, "r", encoding="utf-8") as fp:
        additional_data = json.load(fp=fp)

    matched_keys = set()

    for entry in additional_data:
        measure_id = entry["id"]
        measure_short_title = entry.get("short_title", None)
        if measure_short_title is None:
            print(f"Missing short title: {measure_id}")

        if "state" not in entry:
            print(f"Missing state: {measure_id} - {measure_short_title}")
        else:
            state = entry["state"]
            if state["title"] not in sachstand_keys:
                print(f'Unmatched state title: {measure_id} - {state["title"]}')
            else:
                matched_keys.add(state["title"])

    unmatched_keys = sachstand_keys.difference(matched_keys)

    print(f"Unmatched keys: {unmatched_keys}")


if __name__ == "__main__":
    # extract_from_excel_to_json()
    # fill_implementations_from_additional_data()
    find_unmatched_entries()
